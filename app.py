import streamlit as st
import pandas as pd
import io
import re
import json
import time
from anthropic import Anthropic

# ── CONFIG ────────────────────────────────────────────────────────────────────
CLAUDE_MODEL = "claude-sonnet-4-20250514"
CHUNK_SIZE   = 80

st.set_page_config(
    page_title="RealVal · Rent Roll Standardizer",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── BRANDING & CSS ────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Sora:wght@400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
    background: #f4f6f8;
    color: #0d1f2d;
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 0 2.5rem 4rem; max-width: 1280px; }

.rv-header {
    background: linear-gradient(135deg, #0a2e3d 0%, #0e4a5c 60%, #1a6b7a 100%);
    margin: -1rem -2.5rem 2.5rem;
    padding: 2rem 2.5rem 1.8rem;
    display: flex; align-items: center; justify-content: space-between;
    box-shadow: 0 4px 24px rgba(0,0,0,0.18);
}
.rv-header-left { display: flex; align-items: center; gap: 1.1rem; }
.rv-monogram {
    width: 52px; height: 52px;
    background: linear-gradient(135deg, #2ec4b6, #1a9490);
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-family: 'Sora', sans-serif; font-weight: 700; font-size: 1.35rem;
    color: white; letter-spacing: -0.03em; flex-shrink: 0;
    box-shadow: 0 2px 12px rgba(46,196,182,0.35);
}
.rv-header-text h1 {
    font-family: 'Sora', sans-serif; font-size: 1.45rem; font-weight: 700;
    color: #ffffff; margin: 0; letter-spacing: -0.02em; line-height: 1.2;
}
.rv-header-text p {
    font-size: 0.78rem; color: rgba(255,255,255,0.55);
    margin: 0.2rem 0 0; font-weight: 400;
    letter-spacing: 0.04em; text-transform: uppercase;
}
.rv-header-badge {
    display: flex; align-items: center; gap: 0.5rem;
    background: rgba(46,196,182,0.15); border: 1px solid rgba(46,196,182,0.4);
    border-radius: 20px; padding: 0.4rem 1rem;
    font-size: 0.72rem; font-weight: 600; color: #2ec4b6;
    letter-spacing: 0.06em; text-transform: uppercase;
}

.section-label {
    font-size: 0.68rem; font-weight: 700;
    letter-spacing: 0.12em; text-transform: uppercase; color: #64748b;
    border-bottom: 2px solid #e2e8f0;
    padding-bottom: 0.45rem; margin: 2rem 0 1.1rem;
}

.upload-card {
    background: white; border: 1.5px dashed #cbd5e1;
    border-radius: 10px; padding: 0.5rem 1rem 1rem; transition: border-color 0.2s;
}
.upload-card:hover { border-color: #2ec4b6; }

.stButton > button {
    background: linear-gradient(135deg, #0e4a5c, #1a6b7a) !important;
    color: white !important; font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important; font-size: 0.85rem !important;
    letter-spacing: 0.04em !important; border: none !important;
    border-radius: 8px !important; padding: 0.65rem 2rem !important;
    width: 100% !important; transition: all 0.2s !important;
    box-shadow: 0 2px 8px rgba(14,74,92,0.25) !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #2ec4b6, #1a9490) !important;
    box-shadow: 0 4px 16px rgba(46,196,182,0.35) !important;
    transform: translateY(-1px) !important;
}

.steps-bar {
    display: flex; align-items: center;
    background: white; border-radius: 10px; border: 1px solid #e2e8f0;
    padding: 1rem 1.5rem; margin: 1rem 0 1.5rem; overflow-x: auto;
}
.step { display: flex; align-items: center; gap: 0.5rem; flex-shrink: 0; }
.step-dot {
    width: 28px; height: 28px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 0.72rem; font-weight: 700; flex-shrink: 0;
}
.step-dot.done   { background: #2ec4b6; color: white; }
.step-dot.active { background: #0e4a5c; color: white; box-shadow: 0 0 0 3px rgba(14,74,92,0.15); }
.step-dot.pending{ background: #e2e8f0; color: #94a3b8; }
.step-label { font-size: 0.78rem; font-weight: 500; color: #374151; white-space: nowrap; }
.step-label.pending { color: #94a3b8; }
.step-arrow { color: #cbd5e1; font-size: 0.9rem; margin: 0 0.6rem; flex-shrink: 0; }

.kpi-grid { display: grid; grid-template-columns: repeat(5, 1fr); gap: 0.9rem; margin: 0.5rem 0 1.5rem; }
.kpi-card {
    background: white; border-radius: 10px; border: 1px solid #e2e8f0;
    border-top: 3px solid #0e4a5c; padding: 1rem 1.1rem; transition: box-shadow 0.2s;
}
.kpi-card:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.07); }
.kpi-card.teal  { border-top-color: #2ec4b6; }
.kpi-card.green { border-top-color: #22c55e; }
.kpi-card.amber { border-top-color: #f59e0b; }
.kpi-card.slate { border-top-color: #64748b; }
.kpi-label { font-size: 0.65rem; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; color: #64748b; margin-bottom: 0.35rem; }
.kpi-value { font-family: 'Sora', sans-serif; font-size: 1.7rem; font-weight: 700; color: #0a2e3d; line-height: 1.1; }
.kpi-sub { font-size: 0.7rem; color: #94a3b8; margin-top: 0.2rem; }

.recon-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 0.8rem; margin-bottom: 1.5rem; }
.recon-card {
    background: white; border-radius: 10px; border: 1px solid #e2e8f0;
    padding: 0.9rem 1.1rem; display: flex; align-items: center; gap: 0.9rem;
}
.recon-icon { width: 36px; height: 36px; border-radius: 8px; display: flex; align-items: center; justify-content: center; font-size: 1rem; flex-shrink: 0; }
.recon-icon.pass { background: #dcfce7; }
.recon-icon.warn { background: #fef9c3; }
.recon-icon.fail { background: #fee2e2; }
.recon-body { flex: 1; }
.recon-title  { font-size: 0.78rem; font-weight: 600; color: #0a2e3d; margin-bottom: 0.1rem; }
.recon-detail { font-size: 0.72rem; color: #64748b; }
.recon-badge  { font-size: 0.65rem; font-weight: 700; letter-spacing: 0.06em; text-transform: uppercase; padding: 0.2rem 0.55rem; border-radius: 20px; flex-shrink: 0; }
.badge-pass { background: #dcfce7; color: #166534; }
.badge-warn { background: #fef9c3; color: #854d0e; }
.badge-fail { background: #fee2e2; color: #991b1b; }

.rv-table-wrap { background: white; border-radius: 10px; border: 1px solid #e2e8f0; overflow: hidden; margin-bottom: 1.5rem; }
.rv-table-header {
    background: #0a2e3d; padding: 0.65rem 1rem;
    display: flex; align-items: center; justify-content: space-between;
}
.rv-table-header span { font-size: 0.72rem; font-weight: 600; letter-spacing: 0.08em; text-transform: uppercase; color: rgba(255,255,255,0.7); }
.legend { display: flex; gap: 1.2rem; align-items: center; }
.legend-item { display: flex; align-items: center; gap: 0.35rem; font-size: 0.68rem; color: rgba(255,255,255,0.6); }
.legend-dot { width: 10px; height: 10px; border-radius: 2px; }
.rv-table { width: 100%; border-collapse: collapse; font-size: 0.8rem; }
.rv-table th { background: #f8fafc; color: #374151; font-weight: 600; font-size: 0.72rem; text-align: left; padding: 0.6rem 0.85rem; border-bottom: 2px solid #e2e8f0; letter-spacing: 0.03em; white-space: nowrap; }
.rv-table td { padding: 0.5rem 0.85rem; border-bottom: 1px solid #f1f5f9; color: #1e293b; white-space: nowrap; }
.rv-table tr:last-child td { border-bottom: none; }
.rv-table tr.occupied:hover td { background: #f8fafc; }
.rv-table tr.vacant td  { background: #fffbeb; color: #92400e; }
.rv-table tr.nonrev td  { background: #f0fdf4; color: #166534; font-style: italic; }
.rv-table td.mono  { font-family: 'SF Mono', 'Fira Code', monospace; font-size: 0.75rem; }
.rv-table td.right { text-align: right; }
.rv-table .tag { display: inline-block; font-size: 0.62rem; font-weight: 700; letter-spacing: 0.06em; text-transform: uppercase; padding: 0.15rem 0.5rem; border-radius: 20px; }
.tag-vacant   { background: #fef3c7; color: #92400e; }
.tag-nonrev   { background: #dcfce7; color: #166534; }
.tag-occupied { background: #eff6ff; color: #1d4ed8; }

.stDownloadButton > button {
    background: linear-gradient(135deg, #2ec4b6, #1a9490) !important;
    color: white !important; font-weight: 600 !important;
    font-size: 0.85rem !important; border-radius: 8px !important;
    border: none !important; padding: 0.65rem 2rem !important;
    width: 100% !important; box-shadow: 0 2px 12px rgba(46,196,182,0.3) !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #3dd9ca, #2ec4b6) !important;
    transform: translateY(-1px) !important;
}

.file-pill {
    display: inline-flex; align-items: center; gap: 0.5rem;
    background: #f0fdfa; border: 1px solid #99f6e4;
    border-radius: 20px; padding: 0.35rem 0.9rem;
    font-size: 0.78rem; color: #0f766e; font-weight: 500; margin-bottom: 1rem;
}
.stProgress > div > div > div > div { background: #2ec4b6 !important; }
.rv-footer {
    border-top: 1px solid #e2e8f0; margin-top: 3rem; padding-top: 1rem;
    display: flex; justify-content: space-between; align-items: center;
    font-size: 0.7rem; color: #94a3b8;
}
.rv-footer a { color: #2ec4b6; text-decoration: none; }
</style>
""", unsafe_allow_html=True)


# ── HEADER ────────────────────────────────────────────────────────────────────
def render_header():
    st.markdown("""
    <div class="rv-header">
      <div class="rv-header-left">
        <div class="rv-monogram">RV</div>
        <div class="rv-header-text">
          <h1>Rent Roll Standardizer</h1>
          <p>RealVal &nbsp;·&nbsp; Multifamily Underwriting Intelligence</p>
        </div>
      </div>
      <div class="rv-header-badge">⚡ Powered by Claude AI</div>
    </div>
    """, unsafe_allow_html=True)


# ── PASSWORD GATE ─────────────────────────────────────────────────────────────
def check_password():
    if st.session_state.get("authenticated"):
        return True
    render_header()
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        st.markdown("""
        <div style='background:white;border:1px solid #e2e8f0;border-radius:12px;
        padding:2rem 1.5rem;text-align:center;margin-top:3rem;'>
        <div style='width:52px;height:52px;background:linear-gradient(135deg,#2ec4b6,#1a9490);
        border-radius:10px;display:flex;align-items:center;justify-content:center;
        font-family:Sora,sans-serif;font-weight:700;font-size:1.35rem;color:white;
        margin:0 auto 1rem;'>RV</div>
        <div style='font-family:Sora,sans-serif;font-size:1.2rem;font-weight:700;
        color:#0a2e3d;margin-bottom:0.3rem;'>Internal Access Only</div>
        <div style='font-size:0.82rem;color:#64748b;margin-bottom:1.2rem;'>
        Enter your team password to continue.</div>
        </div>
        """, unsafe_allow_html=True)
        pwd = st.text_input("Password", type="password",
                            label_visibility="collapsed", placeholder="Enter password…")
        if st.button("Sign In"):
            if pwd == st.secrets.get("password", ""):
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    return False


# ── PROGRESS STEPS ────────────────────────────────────────────────────────────
STEPS = ["Reading File", "Sending to Claude", "Building Output", "Ready"]

def render_steps(active: int) -> str:
    html = "<div class='steps-bar'>"
    for i, label in enumerate(STEPS):
        if i < active:
            dot_cls, lbl_cls, icon = "done", "", "✓"
        elif i == active:
            dot_cls, lbl_cls, icon = "active", "", str(i + 1)
        else:
            dot_cls, lbl_cls, icon = "pending", "pending", str(i + 1)
        html += f"""<div class='step'>
          <div class='step-dot {dot_cls}'>{icon}</div>
          <span class='step-label {lbl_cls}'>{label}</span>
        </div>"""
        if i < len(STEPS) - 1:
            html += "<span class='step-arrow'>›</span>"
    html += "</div>"
    return html


# ── CLAUDE CALL ───────────────────────────────────────────────────────────────
def call_claude(client, chunk_csv: str, chunk_num: int, total_chunks: int) -> list:
    prompt = f"""
You are an expert in multifamily real estate underwriting. Standardize the rent roll data below.

REQUIRED OUTPUT COLUMNS:
Unit No | Unit Size (SF) | Market Rent (Monthly) | Effective Rent (Monthly) | Lease Start Date | Lease End Date | Tenant Name

STANDARDIZATION RULES:
1. One row per tenant. Combine all charge lines for the same unit into one row.
2. Effective Rent = the base rent charge ONLY (the primary "rent" line item).
   - ADD housing subsidy if present (charge codes: rentsub, hap, subsidy).
   - SUBTRACT employee discounts (empdisc) and move-in discounts (discnewm) if present.
   - IGNORE and DO NOT ADD any of the following to Effective Rent under any circumstances:
     deposits (security deposit, resident deposit, other deposit, any column with "deposit" in the name),
     utilities, water, electric, gas, trash, pest control,
     pet fees (petfee), parking fees (parkfee), amenity fees (amentfee),
     concessions, late fees, administrative fees, or any other ancillary charges.
   - If you are unsure whether a charge belongs, leave it out. When in doubt, exclude it.
3. If rents are annual, divide by 12. If rent/sf is given, multiply by unit size to get monthly rent.
4. Normalize all dates to MM/DD/YYYY. If date is missing, use null.
5. Vacant rows: include if market rent is shown. Set Effective Rent to null, Tenant Name to "VACANT".
6. Admin/Model units: include them. Set Effective Rent to null. Tenant Name = "ADMIN" or "MODEL".
7. Future/pending leases with no active rent charges: exclude.
8. Remove exact duplicate rows.
9. Do NOT include subtotals, section headers, summary rows, or footers.
10. Round all monetary values to 2 decimal places.

Return ONLY a valid JSON array. No markdown, no explanation, no code fences. Just the raw JSON array starting with [ and ending with ].

Input data (chunk {chunk_num} of {total_chunks}):
{chunk_csv}
"""
    response = client.messages.create(
        model=CLAUDE_MODEL, max_tokens=4096,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = response.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*\n?", "", raw, flags=re.MULTILINE)
    raw = re.sub(r"\n?```\s*$",           "", raw, flags=re.MULTILINE)
    return json.loads(raw.strip())


# ── STANDARDIZE ───────────────────────────────────────────────────────────────
def standardize_rent_roll(df, step_ph, prog_ph, status_ph):
    client     = Anthropic(api_key=st.secrets["anthropic_api_key"])
    total_rows = len(df)
    num_chunks = (total_rows + CHUNK_SIZE - 1) // CHUNK_SIZE
    all_results= []

    step_ph.markdown(render_steps(1), unsafe_allow_html=True)

    for i in range(num_chunks):
        start = i * CHUNK_SIZE
        end   = min((i + 1) * CHUNK_SIZE, total_rows)
        status_ph.markdown(
            f"<small style='color:#64748b;'>Processing chunk {i+1} of {num_chunks} "
            f"({start+1}–{end} of {total_rows} rows)…</small>", unsafe_allow_html=True)

        last_error = None
        for attempt in range(3):
            try:
                rows = call_claude(client, df.iloc[start:end].to_csv(index=False), i+1, num_chunks)
                all_results.extend(rows)
                last_error = None
                break
            except json.JSONDecodeError as e:
                last_error = f"Invalid JSON from Claude (attempt {attempt+1}/3): {e}"
                time.sleep(1)
            except Exception as e:
                last_error = f"Error on chunk {i+1} (attempt {attempt+1}/3): {e}"
                time.sleep(1)

        if last_error:
            status_ph.empty()
            st.error(last_error)
            return pd.DataFrame()

        prog_ph.progress((i + 1) / num_chunks)

    status_ph.empty()
    step_ph.markdown(render_steps(2), unsafe_allow_html=True)

    if not all_results:
        return pd.DataFrame()

    result_df = pd.DataFrame(all_results)
    required  = ["Unit No","Unit Size (SF)","Market Rent (Monthly)",
                 "Effective Rent (Monthly)","Lease Start Date","Lease End Date","Tenant Name"]
    for col in required:
        if col not in result_df.columns:
            result_df[col] = None
    result_df = result_df[required]

    for dc in ["Lease Start Date","Lease End Date"]:
        result_df[dc] = pd.to_datetime(result_df[dc], errors="coerce").dt.strftime("%m/%d/%Y")

    result_df.drop_duplicates(subset=["Unit No"], keep="first", inplace=True)
    result_df.reset_index(drop=True, inplace=True)

    step_ph.markdown(render_steps(3), unsafe_allow_html=True)
    return result_df


# ── COLOR-CODED TABLE ─────────────────────────────────────────────────────────
def render_color_table(df: pd.DataFrame) -> str:
    COLS = ["Unit No","Unit Size (SF)","Market Rent (Monthly)",
            "Effective Rent (Monthly)","Lease Start Date","Lease End Date","Tenant Name"]
    hdr = "".join(f"<th>{c}</th>" for c in COLS)
    body = ""
    for _, row in df.iterrows():
        name = str(row.get("Tenant Name","") or "").strip().upper()
        if name == "VACANT":
            tr, tag = "vacant", "<span class='tag tag-vacant'>Vacant</span>"
        elif name in ("ADMIN","MODEL"):
            tr, tag = "nonrev", f"<span class='tag tag-nonrev'>{name.title()}</span>"
        else:
            tr, tag = "occupied","<span class='tag tag-occupied'>Occupied</span>"

        def fmt(col, val):
            if val is None or (isinstance(val,float) and pd.isna(val)): return "—"
            if col in ("Market Rent (Monthly)","Effective Rent (Monthly)"):
                try: return f"${float(val):,.2f}"
                except: return str(val)
            if col == "Unit Size (SF)":
                try: return f"{int(float(val)):,}"
                except: return str(val)
            return str(val)

        cells = ""
        for col in COLS:
            val = row.get(col)
            cls = ("mono " if col=="Unit No" else "") + \
                  ("right " if col in ("Market Rent (Monthly)","Effective Rent (Monthly)","Unit Size (SF)") else "")
            display = tag if col == "Tenant Name" else fmt(col, val)
            cells += f"<td class='{cls}'>{display}</td>"
        body += f"<tr class='{tr}'>{cells}</tr>"

    return f"""
    <div class='rv-table-wrap'>
      <div class='rv-table-header'>
        <span>Standardized Output — {len(df)} rows</span>
        <div class='legend'>
          <div class='legend-item'><div class='legend-dot' style='background:#fef3c7;border:1px solid #fcd34d;'></div>Vacant</div>
          <div class='legend-item'><div class='legend-dot' style='background:#dcfce7;border:1px solid #86efac;'></div>Admin / Model</div>
          <div class='legend-item'><div class='legend-dot' style='background:#eff6ff;border:1px solid #bfdbfe;'></div>Occupied</div>
        </div>
      </div>
      <div style='overflow-x:auto;max-height:480px;overflow-y:auto;'>
        <table class='rv-table'><thead><tr>{hdr}</tr></thead><tbody>{body}</tbody></table>
      </div>
    </div>"""


# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
def build_excel(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Standardized Rent Roll"
    navy="0A2E3D"; thin=Side(style="thin",color="E2E8F0"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    headers=["Unit No","Unit Size (SF)","Market Rent (Monthly)","Effective Rent (Monthly)","Lease Start Date","Lease End Date","Tenant Name"]
    for col,h in enumerate(headers,1):
        c=ws.cell(row=1,column=col,value=h)
        c.font=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
        c.fill=PatternFill("solid",start_color=navy)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=bdr
    ws.row_dimensions[1].height=28
    for row_idx,rec in enumerate(df.itertuples(index=False),2):
        name=str(rec[6] or "").upper()
        is_vac=name=="VACANT"; is_nr=name in ("ADMIN","MODEL")
        fill=(PatternFill("solid",start_color="FFFBEB") if is_vac else
              PatternFill("solid",start_color="F0FDF4") if is_nr  else
              PatternFill("solid",start_color="FFFFFF") if row_idx%2 else
              PatternFill("solid",start_color="F8FAFC"))
        for col,val in enumerate(rec,1):
            c=ws.cell(row=row_idx,column=col,value=val)
            c.border=bdr; c.fill=fill
            c.font=Font(name="Calibri",size=10,italic=(is_vac or is_nr),
                        color=("92400E" if is_vac else "166534" if is_nr else "1E293B"))
            if col==2: c.number_format="#,##0"; c.alignment=Alignment(horizontal="right")
            elif col in (3,4): c.number_format="$#,##0.00"; c.alignment=Alignment(horizontal="right")
            elif col in (5,6): c.alignment=Alignment(horizontal="center")
            else: c.alignment=Alignment(horizontal="left")
    last=len(df)+1; tot=last+1
    for col in range(1,8):
        c=ws.cell(row=tot,column=col)
        c.fill=PatternFill("solid",start_color=navy)
        c.font=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
        c.border=bdr; c.alignment=Alignment(horizontal="right")
    ws.cell(row=tot,column=1,value="TOTALS / AVERAGES").alignment=Alignment(horizontal="left")
    ws.cell(row=tot,column=2,value=f"=SUM(B2:B{last})").number_format="#,##0"
    ws.cell(row=tot,column=3,value=f"=AVERAGE(C2:C{last})").number_format="$#,##0.00"
    ws.cell(row=tot,column=4,value=f'=AVERAGEIF(D2:D{last},"<>",D2:D{last})').number_format="$#,##0.00"
    nr=tot+2
    for i,note in enumerate(["Notes:","• Yellow rows = Vacant units","• Green rows = Admin / Model units",
        "• Effective Rent = base rent + subsidy − employee/move-in discounts",
        "• Deposits, utilities, pet/parking fees excluded from Effective Rent"]):
        c=ws.cell(row=nr+i,column=1,value=note)
        c.font=Font(name="Calibri",bold=(i==0),size=9,
                    color=("166534" if "Green" in note else "92400E" if "Yellow" in note else "595959"))
    for i,w in enumerate([14,14,22,24,16,16,26],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:G{last}"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ── RECONCILIATION ────────────────────────────────────────────────────────────
def render_recon(df, orig_rows):
    occupied   = df[~df["Tenant Name"].str.upper().isin(["VACANT","ADMIN","MODEL"]) & df["Tenant Name"].notna()]
    vacant     = df[df["Tenant Name"].str.upper()=="VACANT"]
    clean_rows = len(df)
    clean_mkt  = pd.to_numeric(df["Market Rent (Monthly)"],errors="coerce").sum()
    missing    = [c for c in ["Unit No","Unit Size (SF)","Market Rent (Monthly)",
                              "Effective Rent (Monthly)","Lease Start Date","Lease End Date","Tenant Name"]
                  if c not in df.columns]
    size_warn  = clean_rows < orig_rows * 0.3
    checks = [
        {"title":"All Required Columns Present",
         "detail":"7 of 7 columns found" if not missing else f"Missing: {', '.join(missing)}",
         "status":"pass" if not missing else "fail","icon":"✓" if not missing else "✗","badge":"Pass" if not missing else "Fail"},
        {"title":"Row Count",
         "detail":f"{clean_rows} cleaned rows from {orig_rows} raw rows",
         "status":"warn" if size_warn else "pass","icon":"⚠" if size_warn else "✓","badge":"Review" if size_warn else "Pass"},
        {"title":"Market Rent Populated",
         "detail":f"Total market rent: ${clean_mkt:,.2f}",
         "status":"pass" if clean_mkt>0 else "fail","icon":"✓" if clean_mkt>0 else "✗","badge":"Pass" if clean_mkt>0 else "Fail"},
        {"title":"Occupied Tenant Count",
         "detail":f"{len(occupied)} occupied · {len(vacant)} vacant",
         "status":"pass" if len(occupied)>0 else "warn","icon":"✓" if len(occupied)>0 else "⚠","badge":"Pass" if len(occupied)>0 else "Review"},
    ]
    html = "<div class='recon-grid'>"
    for c in checks:
        html += f"""<div class='recon-card'>
          <div class='recon-icon {c["status"]}'>{c["icon"]}</div>
          <div class='recon-body'>
            <div class='recon-title'>{c["title"]}</div>
            <div class='recon-detail'>{c["detail"]}</div>
          </div>
          <span class='recon-badge badge-{c["status"]}'>{c["badge"]}</span>
        </div>"""
    html += "</div>"
    return html


# ── KPI SUMMARY ───────────────────────────────────────────────────────────────
def render_kpis(df) -> str:
    occupied   = df[~df["Tenant Name"].str.upper().isin(["VACANT","ADMIN","MODEL"]) & df["Tenant Name"].notna()]
    vacant     = df[df["Tenant Name"].str.upper()=="VACANT"]
    total      = len(df)
    occ_pct    = (len(occupied)/total*100) if total else 0
    avg_mkt    = pd.to_numeric(df["Market Rent (Monthly)"],errors="coerce").mean()
    avg_eff    = pd.to_numeric(occupied["Effective Rent (Monthly)"],errors="coerce").mean()
    tpgi       = pd.to_numeric(df["Market Rent (Monthly)"],errors="coerce").sum()
    def fc(v): return f"${v:,.0f}" if pd.notna(v) and v>0 else "—"
    cards = [
        ("Total Units",       f"{total}",         "All incl. vacant & non-rev",""),
        ("Occupancy Rate",    f"{occ_pct:.1f}%",  f"{len(occupied)} occ / {len(vacant)} vac","teal"),
        ("Avg Market Rent",   fc(avg_mkt),         "Monthly, all units","slate"),
        ("Avg Effective Rent",fc(avg_eff),          "Occupied units only","green"),
        ("Total Potential GI",fc(tpgi),             "Sum of all market rents","amber"),
    ]
    html = "<div class='kpi-grid'>"
    for label,value,sub,cls in cards:
        html += f"""<div class='kpi-card {cls}'>
          <div class='kpi-label'>{label}</div>
          <div class='kpi-value'>{value}</div>
          <div class='kpi-sub'>{sub}</div>
        </div>"""
    html += "</div>"
    return html


# ── MAIN ──────────────────────────────────────────────────────────────────────
if not check_password():
    st.stop()

render_header()

st.markdown("<div class='section-label'>Upload Rent Roll</div>", unsafe_allow_html=True)
col_up, col_info = st.columns([2, 1])
with col_up:
    st.markdown("<div class='upload-card'>", unsafe_allow_html=True)
    uploaded_file = st.file_uploader("Drop an Excel file here or click to browse", type=["xlsx"])
    st.markdown("</div>", unsafe_allow_html=True)
with col_info:
    st.markdown("""
    <div style='background:white;border:1px solid #e2e8f0;border-radius:10px;padding:1rem 1.2rem;font-size:0.8rem;color:#64748b;line-height:1.8;'>
    <strong style='color:#0a2e3d;display:block;margin-bottom:0.4rem;'>Accepted formats</strong>
    Any broker Excel layout · Multi-line tenants<br>Annual or monthly rents · Rent-per-SF<br><br>
    <strong style='color:#0a2e3d;display:block;margin-bottom:0.4rem;'>Output</strong>
    One row per tenant · Color-coded preview<br>Downloadable .xlsx with totals row
    </div>""", unsafe_allow_html=True)

if uploaded_file:
    st.markdown(f"<div class='file-pill'>📄 {uploaded_file.name} &nbsp;·&nbsp; {uploaded_file.size/1024:.1f} KB</div>", unsafe_allow_html=True)
    col_btn, _ = st.columns([1, 2])
    with col_btn:
        run = st.button("⚡  Standardize Rent Roll")

    if run:
        st.markdown("<div class='section-label'>Processing</div>", unsafe_allow_html=True)
        step_ph   = st.empty()
        step_ph.markdown(render_steps(0), unsafe_allow_html=True)
        prog_ph   = st.progress(0)
        status_ph = st.empty()

        try:
            original_df = pd.read_excel(uploaded_file, header=None)
        except Exception as e:
            st.error(f"Could not read file: {e}"); st.stop()

        status_ph.markdown(f"<small style='color:#64748b;'>File loaded — {len(original_df)} raw rows detected</small>", unsafe_allow_html=True)
        time.sleep(0.4)

        standardized_df = standardize_rent_roll(original_df, step_ph, prog_ph, status_ph)

        if standardized_df.empty:
            st.error("Standardization failed or returned no data. Please check the file and try again.")
            st.stop()

        prog_ph.empty(); status_ph.empty()

        st.markdown("<div class='section-label'>Underwriting Summary</div>", unsafe_allow_html=True)
        st.markdown(render_kpis(standardized_df), unsafe_allow_html=True)

        st.markdown("<div class='section-label'>Reconciliation Checks</div>", unsafe_allow_html=True)
        st.markdown(render_recon(standardized_df, len(original_df)), unsafe_allow_html=True)

        st.markdown("<div class='section-label'>Preview</div>", unsafe_allow_html=True)
        st.markdown(render_color_table(standardized_df), unsafe_allow_html=True)

        st.markdown("<div class='section-label'>Download</div>", unsafe_allow_html=True)
        from datetime import date
        today = date.today().strftime("%m%d%Y")
        base  = uploaded_file.name.replace(".xlsx","")
        col_dl, _ = st.columns([1, 2])
        with col_dl:
            st.download_button(
                label="⬇  Download Standardized Rent Roll",
                data=build_excel(standardized_df),
                file_name=f"{base}_standardized_{today}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
else:
    st.info("Upload an Excel rent roll above to get started.")

st.markdown("""
<div class='rv-footer'>
  <span>© 2026 <a href='https://therealval.com' target='_blank'>RealVal</a> · Internal Tool · All Rights Reserved</span>
  <span>Powered by Claude AI · For analyst use only</span>
</div>""", unsafe_allow_html=True)
